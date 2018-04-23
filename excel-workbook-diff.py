#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl
from difflib import SequenceMatcher
from difflib import Differ
from datetime import datetime

def diff_excel_workbook(file1, file2, title_row=1, start_row=2, sheet_index=0):
    sheet1 = read_workbook(file1, sheet_index)
    sheet2 = read_workbook(file2, sheet_index)
    
    sheet1_header_titles = []
    sheet2_header_titles = []
    for cells in sheet1.iter_cols(min_row=title_row, max_row=title_row):
        sheet1_header_titles.append(cell_to_text_oneline(cells[0]))
    for cells in sheet2.iter_cols(min_row=title_row, max_row=title_row):
        sheet2_header_titles.append(cell_to_text_oneline(cells[0]))
    if sheet1_header_titles != sheet2_header_titles:
        raise Exception("title fields not matched. {} vs {}", sheet1_header_titles, sheet2_header_titles)
    
    table1 = read_sheet_table(sheet1, start_row=start_row)
    table2 = read_sheet_table(sheet2, start_row=start_row)

    diff_result = diff_two_tables(table1, table2)
    return diff_result, sheet1_header_titles


def format_diff_two_tables(diff_result, header_titles, primary_key_columns, prefix_row="=== ", prefix_column="# "):
    format_result = ""
    for opcode, field1, field2 in diff_result:
        if opcode == "insert":
            format_result += prefix_row + " ".join([field1[i - 1] for i in primary_key_columns]) + "\n"
            for i in range(0, len(field1)):
                format_result += prefix_column + header_titles[i] + "\n"
                format_result += add_prefix_each_line(field1[i], "+") + "\n"
        elif opcode == "delete":
            format_result += prefix_row + " ".join([field1[i - 1] for i in primary_key_columns]) + "\n"
            for i in range(0, len(field1)):
                format_result += prefix_column + header_titles[i] + "\n"
                format_result += add_prefix_each_line(field1[i], "-") + "\n"
        elif opcode == "replace":
            format_result += prefix_row + " ".join([field2[i - 1] for i in primary_key_columns]) + "\n"
            for i in range(0, len(field1)):
                if field1[i] == field2[i]:
                    continue
                format_result += prefix_column + header_titles[i] + "\n"
                field1array = field1[i].splitlines()
                field2array = field2[i].splitlines()
                match = SequenceMatcher(None, field1array, field2array)
                for tag, i1, i2, j1, j2 in match.get_opcodes():
                    if tag == "equal":
                        for line in field1array[i1:i2]:
                            format_result += " " + line + "\n"
                    if tag == "delete" or tag == "replace":
                        for line in field1array[i1:i2]:
                            format_result += "-" + line + "\n"
                    if tag == "insert" or tag == "replace":
                        for line in field2array[j1:j2]:
                            format_result += "+" + line + "\n"
            for i in range(len(field1), len(field2)):
                format_result += "+" + field2[i] + "\n"
    return format_result

def diff_two_tables(table1, table2):
    matcher = SequenceMatcher(None, table1, table2)
    result = []
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == "equal":
            pass
        elif tag == "insert":
            for i in range(j1, j2):
                result.append(("insert", table2[i], None,))
        elif tag == "delete":
            for i in range(i1, i2):
                result.append(("delete", table1[i], None,))
        elif tag == "replace":
            insert_entries = list(table2[j1:j2])
            for i in range(i1, i2):
                while True:
                    replace_one_line_score = 0
                    insert_one_line_score = 0
                    delete_one_line_score = 0
                    if len(insert_entries) >= 1:
                        replace_one_line_score = count_exact_entries_in_tuple(table1[i], insert_entries[0])
                    if len(insert_entries) >= 2:
                        insert_one_line_score = count_exact_entries_in_tuple(table1[i], insert_entries[1])
                    if i < i2 - 1 and len(insert_entries) >= 1:
                        delete_one_line_score = count_exact_entries_in_tuple(table1[i + 1], insert_entries[0])
                    if replace_one_line_score == 0 and insert_one_line_score == 0 and delete_one_line_score == 0:
                        result.append(("delete", table1[i], None,))
                        break
                    elif max(replace_one_line_score, insert_one_line_score, delete_one_line_score) == replace_one_line_score:
                        result.append(("replace", table1[i], insert_entries.pop(0)))
                        break
                    elif max(replace_one_line_score, insert_one_line_score, delete_one_line_score) == insert_one_line_score:
                        result.append(("insert", insert_entries.pop(0), None,))
                    else:
                        result.append(("delete", table1[i], None,))
                        break
            for entry in insert_entries:
                result.append(("insert", entry, None,))
    return result
        
def count_exact_entries_in_tuple(tuple1, tuple2):
    count = 0
    for i in range(min(len(tuple1), len(tuple2))):
        if tuple1[i] == tuple2[i]:
            count += 1
    return count

def read_workbook(filename, sheet_index):
    workbook = openpyxl.load_workbook(filename, data_only=True)
    if str(sheet_index).isdigit():
        worksheet = workbook.worksheets[sheet_index]
    else:
        worksheet = workbook.get_sheet_by_name(sheet_index)
    return worksheet

def read_sheet_table(sheet, start_row=2):
    row_index = 0
    table = []
    for row in sheet.rows:
        row_index += 1
        if row_index < start_row:
            continue
        line = []
        for cell in row:
            text = cell_to_text_multiline(cell.value)
            line.append(str(text))
        table.append(tuple(line))
    return table

def cell_to_text_oneline(cell):
    text = cell.value
    if text is None:
        text = ""
    while True:
        if text == "":
            return ""
        if text[-1] == "\n":
            text = text[0:-1]
            continue
        break
    text = text.replace("\n", " ")
    return text

def cell_to_text_multiline(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value == value.replace(hour=0, minute=0, second=0, microsecond=0):
            return datetime.date(value)
        return value

    text = str(value)
    while True:
        if text == "":
            return ""
        if text[-1] == "\n":
            text = text[0:-1]
            continue
        break
    return text

def add_prefix_each_line(text, prefix):
    text = prefix + text.replace("\n", "\n" + prefix)
    return text

if __name__ == '__main__':
    import sys
    import argparse
    parser = argparse.ArgumentParser(add_help=True)
    parser.add_argument("excelfile1")
    parser.add_argument("excelfile2")
    parser.add_argument("--title-row", metavar="1", default=1, type=int,
        help="row number of heading row")
    parser.add_argument("--table-start-row", metavar="2", default=2, type=int,
        help="first row number of table data, excluding headings")
    parser.add_argument("--row-heading-prefix", metavar='"=== "', default="=== ",
        help="prefix that displays at each row")
    parser.add_argument("--row-heading-display-cols", metavar='1,2', default="1,2",
        help="column number to display each row")
    parser.add_argument("--column-heading-prefix", metavar='"# "', default="# ",
        help="prefix that displays at each column")
    arg = parser.parse_args()
    file1 = sys.argv[1]
    file2 = sys.argv[2]
    row_heading_display_cols = tuple(map(lambda n: int(n), arg.row_heading_display_cols.split(",")))
    
    diff_result, header_titles = diff_excel_workbook(arg.excelfile1, arg.excelfile2, title_row=int(arg.title_row), start_row=int(arg.table_start_row))
    text = format_diff_two_tables(diff_result, header_titles, row_heading_display_cols, prefix_row=arg.row_heading_prefix, prefix_column=arg.column_heading_prefix)
    print(text)
